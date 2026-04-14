from openpyxl import load_workbook
from Work_Here.create_workbook.data import names, months, days


def createbook(filename, month_num, year):
    if days[month_num - 1] == 0:
        wb = load_workbook("D:\\Billing\\Work_Here\\model_excel\\model_0.xlsx")
        ws = wb.active
        #this line
        template_images = ws._images.copy() if hasattr(ws, '_images') else []
        for i in names:
            title_number = 0
            for products in names[i]:
                new_sheet = wb.copy_worksheet(ws)
                new_sheet["C6"] = i
                if title_number == 0:
                    new_sheet.title = i
                else:
                    new_sheet.title = f"{i} ({title_number})"
                new_sheet["B7"] = f"மாதம்: {months[month_num - 1]} {year} "
                new_sheet["E7"] = f"பொருள்: {products}"

                # Copy images to the new sheet
                for img in template_images:
                    # Create a copy of each image with the same properties
                    if hasattr(img, 'path') and img.path:
                        copied_img = Image(img.path)
                        copied_img.anchor = img.anchor
                        new_sheet.add_image(copied_img)

                title_number += 1
        wb.save(f"D:\\Billing\\Work_Here\\montly sheets\\{filename}.xlsx")

    elif days[month_num - 1] == 1:
        wb = load_workbook("D:\\Billing\\Work_Here\\model_excel\\model_1.xlsx")
        ws = wb.active
        template_images = ws._images.copy() if hasattr(ws, '_images') else []
        for i in names:
            title_number = 0
            for products in names[i]:
                new_sheet = wb.copy_worksheet(ws)
                new_sheet["C6"] = i
                if title_number == 0:
                    new_sheet.title = i
                else:
                    new_sheet.title = f"{i} ({title_number})"
                new_sheet["B7"] = f"மாதம்: {months[month_num - 1]} {year} "
                new_sheet["E7"] = f"பொருள்: {products}"

                # Copy images to the new sheet
                for img in template_images:
                    # Create a copy of each image with the same properties
                    if hasattr(img, 'path') and img.path:
                        copied_img = Image(img.path)
                        copied_img.anchor = img.anchor
                        new_sheet.add_image(copied_img)

                title_number += 1
        wb.save(f"D:\\Billing\\Work_Here\\montly sheets\\{filename}.xlsx")

    elif days[month_num - 1] == 3:
        wb = load_workbook("D:\\Billing\\Work_Here\\model_excel\\model_3.xlsx")
        ws = wb.active
        template_images = ws._images.copy() if hasattr(ws, '_images') else []
        for i in names:
            title_number = 0
            for products in names[i]:
                new_sheet = wb.copy_worksheet(ws)
                new_sheet["C6"] = i
                if title_number == 0:
                    new_sheet.title = i
                else:
                    new_sheet.title = f"{i} ({title_number})"
                new_sheet["B7"] = f"மாதம்: {months[month_num - 1]} {year} "
                new_sheet["E7"] = f"பொருள்: {products}"

                # Copy images to the new sheet
                for img in template_images:
                    # Create a copy of each image with the same properties
                    if hasattr(img, 'path') and img.path:
                        copied_img = Image(img.path)
                        copied_img.anchor = img.anchor
                        new_sheet.add_image(copied_img)

                title_number += 1
        wb.save(f"D:\\Billing\\Work_Here\\montly sheets\\{filename}.xlsx")


