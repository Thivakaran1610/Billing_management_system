from openpyxl import load_workbook
from Work_Here.create_workbook.data import names, name_list


def editsheet(path, book_code, variety, date, wei_val1, pri_val2):
    wb = load_workbook(f"D:\Billing\Work_Here\montly sheets\{path}")
    book = f"{name_list[book_code]} ({names[name_list[book_code]].index(variety)})"
    ws = wb[book]

    wei_cell = f"C{str(date + 9)}"
    pri_cell = f"D{str(date + 9)}"
    ws[wei_cell] = wei_val1
    ws[pri_cell] = pri_val2
    wb.save(f"D:\Billing\Work_Here\montly sheets\{path}")


def list_type(book_num):
    return names[name_list[book_num]]

