from tkinter import Tk, Button, W, CENTER, END
import tkinter.ttk as ttk
from openpyxl import load_workbook


class ExcelRows(Tk):

    def __init__(self, ):
        super().__init__()
        self.initialize_display()

    def initialize_display(self):
        # Configure root object
        self.title('Excel Rows')
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(0, weight=1)
        self.config(background="yellow")

        # Load Excel Workbook and Sheet
        self.wb = load_workbook("D:\\Billing\\Work_Here\\montly sheets\\JulyBill.xlsx")
        self.ws = self.wb.active

        # Initialise the row count
        self.cur_row = 7

        # Define 'Previous' button
        self.button_prev = Button(self, text='Get prev Row',
                                  command=lambda: self.insert_row_data('prev'))
        self.button_prev.grid(row=5, column=0, sticky=W)

        # Define 'Next' button
        self.button_next = Button(self, text='Get next Row',
                                  command=lambda: self.insert_row_data('next'))
        self.button_next.grid(row=5, column=1, sticky=W)

        # Define 'Exit' button
        self.exit_button = Button(self, text="Exit", command=self.quit)
        self.exit_button.grid(row=5, column=2)

        # Define treeview display
        col_list = self.get_columns()
        self.treeview = ttk.Treeview(self, columns=col_list, height=3)

        # Set the headings and column widths
        for enum, col_header in enumerate(col_list):
            self.treeview.heading(enum, text=col_header)
            self.treeview.column(col_header, width=50)
        self.treeview.heading('#0', text='Row', anchor=CENTER)
        self.treeview.column("#0", width=65)
        self.treeview.grid(row=1, columnspan=3, sticky='w')

    def get_columns(self):
        # Get a list of headers from row 1 and enter as Treeview Headers
        header_names = []
        for cell in self.ws[7]:
            header_names.append(cell.value)
        return header_names

    def insert_row_data(self, seq):
        # Get the row as selected by the 'next' or 'prev' button and insert to Treeview
        row_list = []

        # Clear existing Treeview data before writing new data to same row
        for item in self.treeview.get_children():
            self.treeview.delete(item)

        # 'next' button increments the current row, if the end of the used rows is reached
        # the row is set back to the start and loops
        if seq == 'next':
            self.cur_row += 1
            if self.cur_row == self.ws.max_row:
                self.cur_row = self.ws.min_row

        # 'prev' button decrements the current row, if the start of the used rows is reached
        # the row is set to the last used row and loops
        elif seq == 'prev':
            self.cur_row -= 1
            if self.cur_row == self.ws.min_row - 1 or self.cur_row == -1:
                self.cur_row = self.ws.max_row - 1

        # Builds a list of cell values in the set row
        for row_cells in self.ws.iter_cols(min_col=self.ws.min_column,
                                           max_col=self.ws.max_column):
            row_list.append(str(row_cells[self.cur_row].value))

        # Insert the row_list into the Treeview
        self.treeview.insert('', END,
                             text="Row " + str(self.cur_row),
                             values=[row_list][0])




app = ExcelRows()
app.mainloop()
